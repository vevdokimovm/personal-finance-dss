"""Экспорт плана распределения в XLSX и PDF (P2.5).

Тот же контент, что у CSV-экспорта (`/api/planning/export.csv`): показатели Rt/Lt/Dt/BLR,
рекомендованное распределение и топ-3 альтернатив. PDF переиспользует шрифт и табличный
стиль из `report_pdf` (Liberation Sans — кириллица работает на любой системе; суммы — через
`_money` с « руб», без символа ₽, чтобы не зависеть от наличия глифа). Возвращают байты.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from app.services.report_pdf import _BRAND, _FONT, _FONT_BOLD, _ensure_fonts, _money, _table


def _f(value: Any, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _extract(result: dict[str, Any]) -> dict[str, Any]:
    """Приводит результат _compute_plan к плоской структуре для обоих форматов."""
    ind = result.get("indicators", {}) or {}
    top3 = result.get("top3", []) or []
    best = top3[0] if top3 else {}
    return {
        "profile": result.get("risk_profile", ""),
        "best_name": best.get("name", ""),
        "rt": _f(ind.get("Rt")),
        "lt": _f(ind.get("Lt")),
        "dt_pct": _f(ind.get("Dt")) * 100,
        "blr": _f(ind.get("BLR")),
        "obl": _f(best.get("x_obligations")),
        "res": _f(best.get("x_reserve")),
        "goals": _f(best.get("x_goals")),
        "utility": _f(best.get("utility")),
        "top3": [
            {
                "i": i,
                "name": a.get("name", ""),
                "obl": _f(a.get("x_obligations")),
                "res": _f(a.get("x_reserve")),
                "goals": _f(a.get("x_goals")),
                "u": _f(a.get("utility")),
            }
            for i, a in enumerate(top3, start=1)
        ],
    }


# ── XLSX ──────────────────────────────────────────────────────────────────
_MONEY = '#,##0" ₽"'
_PCT = '0.0'
_NUM2 = '0.00'
_U = '0.000'


def plan_to_xlsx(result: dict[str, Any]) -> bytes:
    """Собирает XLSX плана распределения и возвращает байты."""
    d = _extract(result)
    wb = Workbook()
    ws = wb.active
    ws.title = "План"

    head_font = Font(name="Arial", bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2BBF6A")
    base = Font(name="Arial")
    bold = Font(name="Arial", bold=True)

    def header(cells: list[str]) -> None:
        ws.append(cells)
        r = ws.max_row
        for col in range(1, len(cells) + 1):
            ws.cell(r, col).font = head_font
            ws.cell(r, col).fill = head_fill

    def kv(label: str, value: Any, fmt: str | None = None) -> None:
        ws.append([label, value])
        r = ws.max_row
        ws.cell(r, 1).font = base
        c = ws.cell(r, 2)
        c.font = base
        if fmt:
            c.number_format = fmt

    ws.append(["FINPILOT — план распределения"])
    ws.cell(1, 1).font = Font(name="Arial", bold=True, size=14, color="2BBF6A")
    ws.append(["Профиль риска", d["profile"]])
    ws.cell(2, 1).font = bold
    ws.append([])

    header(["ПОКАЗАТЕЛЬ", "ЗНАЧЕНИЕ"])
    kv("Свободные деньги (Rt), ₽", round(d["rt"]), _MONEY)
    kv("Ликвидность (Lt)", d["lt"], _NUM2)
    kv("Долговая нагрузка (Dt), %", d["dt_pct"], _PCT)
    kv("Подушка (BLR), мес", d["blr"], _NUM2)
    ws.append([])

    header(["РЕКОМЕНДОВАННОЕ РАСПРЕДЕЛЕНИЕ", d["best_name"]])
    kv("На досрочное погашение, ₽", round(d["obl"]), _MONEY)
    kv("В подушку безопасности, ₽", round(d["res"]), _MONEY)
    kv("На цели, ₽", round(d["goals"]), _MONEY)
    kv("Оценка полезности U", d["utility"], _U)
    ws.append([])

    header(["#", "Название", "Долг", "Резерв", "Цели", "Оценка"])
    for a in d["top3"]:
        ws.append([a["i"], a["name"], round(a["obl"]), round(a["res"]), round(a["goals"]), a["u"]])
        r = ws.max_row
        for col in (3, 4, 5):
            ws.cell(r, col).number_format = _MONEY
        ws.cell(r, 6).number_format = _U

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24
    for col in ("C", "D", "E", "F"):
        ws.column_dimensions[col].width = 13

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ── PDF ───────────────────────────────────────────────────────────────────
def plan_to_pdf(result: dict[str, Any]) -> bytes:
    """Собирает PDF плана распределения и возвращает байты (кириллица — Liberation Sans)."""
    d = _extract(result)
    _ensure_fonts()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        topMargin=20 * mm, bottomMargin=18 * mm, leftMargin=18 * mm, rightMargin=18 * mm,
        title="FINPILOT — план распределения",
    )
    h1 = ParagraphStyle("H1", fontName=_FONT_BOLD, fontSize=18, leading=22, textColor=_BRAND)
    h2 = ParagraphStyle("H2", fontName=_FONT_BOLD, fontSize=13, leading=18, spaceBefore=12, spaceAfter=6)
    small = ParagraphStyle("Small", fontName=_FONT, fontSize=8, textColor=colors.grey, leading=11)

    story: list = [
        Paragraph("FINPILOT — план распределения", h1),
        Paragraph(f"Профиль риска: {d['profile']}", small),
        Spacer(1, 12),
        Paragraph("Показатели", h2),
        _table([
            ["Показатель", "Значение"],
            ["Свободные деньги (Rt)", _money(d["rt"])],
            ["Ликвидность (Lt)", f"{d['lt']:.2f}"],
            ["Долговая нагрузка (Dt)", f"{d['dt_pct']:.1f}%"],
            ["Подушка (BLR), мес", f"{d['blr']:.2f}"],
        ]),
        Spacer(1, 8),
        Paragraph("Рекомендованное распределение", h2),
        _table([
            ["Направление", "Значение"],
            ["Лучший вариант", d["best_name"]],
            ["На досрочное погашение", _money(d["obl"])],
            ["В подушку безопасности", _money(d["res"])],
            ["На цели", _money(d["goals"])],
            ["Оценка полезности U", f"{d['utility']:.3f}"],
        ]),
        Spacer(1, 8),
        Paragraph("Все варианты (топ-3)", h2),
    ]
    rows = [["#", "Название", "Долг", "Резерв", "Цели", "Оценка"]]
    for a in d["top3"]:
        rows.append([
            str(a["i"]), a["name"], _money(a["obl"]), _money(a["res"]),
            _money(a["goals"]), f"{a['u']:.3f}",
        ])
    story.append(_table(rows))
    story.append(Spacer(1, 16))
    story.append(Paragraph(
        "Информационный документ. Не является индивидуальной инвестиционной рекомендацией "
        "в значении Федерального закона № 39-ФЗ.", small,
    ))

    doc.build(story)
    return buf.getvalue()
