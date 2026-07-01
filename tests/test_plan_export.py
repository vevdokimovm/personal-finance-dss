"""Юнит-тесты экспорта плана в XLSX и PDF (P2.5)."""
import io

import openpyxl

from app.services.plan_export import _extract, plan_to_pdf, plan_to_xlsx

_RESULT = {
    "risk_profile": "3 Сбалансированный",
    "indicators": {"Rt": 39500, "Lt": 0.405, "Dt": 0.347, "BLR": 1.2},
    "top3": [
        {"name": "Всё в резерв", "x_obligations": 0, "x_reserve": 39500, "x_goals": 0, "utility": 0.805},  # noqa: E501
        {"name": "Долг+резерв", "x_obligations": 10000,
            "x_reserve": 29500, "x_goals": 0, "utility": 0.78},
        {"name": "Баланс", "x_obligations": 13000, "x_reserve": 13000, "x_goals": 13500, "utility": 0.75},  # noqa: E501
    ],
}


class TestExtract:
    def test_flattens_indicators_and_best(self):
        d = _extract(_RESULT)
        assert d["profile"] == "3 Сбалансированный"
        assert d["rt"] == 39500.0
        assert round(d["dt_pct"], 1) == 34.7  # доля → проценты
        assert d["best_name"] == "Всё в резерв"
        assert len(d["top3"]) == 3

    def test_empty_result_safe(self):
        d = _extract({})
        assert d["rt"] == 0.0
        assert d["top3"] == []
        assert d["best_name"] == ""


class TestXlsx:
    def test_returns_valid_xlsx_bytes(self):
        data = plan_to_xlsx(_RESULT)
        assert isinstance(data, bytes)
        assert data[:4] == b"PK\x03\x04"

    def test_content_roundtrips(self):
        wb = openpyxl.load_workbook(io.BytesIO(plan_to_xlsx(_RESULT)))
        rows = [[c.value for c in r] for r in wb.active.iter_rows()]
        assert rows[0][0] == "FINPILOT — план распределения"
        # кириллица в профиле сохранилась
        assert any(r[1] == "3 Сбалансированный" for r in rows if r[0] == "Профиль риска")
        # значение Rt записано числом (для форматирования в Excel)
        rt = next(r[1] for r in rows if r[0] and "Rt" in str(r[0]) and "Свободные" in str(r[0]))
        assert rt == 39500

    def test_top3_rows_present(self):
        wb = openpyxl.load_workbook(io.BytesIO(plan_to_xlsx(_RESULT)))
        names = [c.value for col in wb.active.iter_cols(min_col=2, max_col=2) for c in col]
        assert "Баланс" in names

    def test_empty_result_does_not_crash(self):
        assert plan_to_xlsx({})[:4] == b"PK\x03\x04"


class TestPdf:
    def test_returns_valid_pdf_bytes(self):
        data = plan_to_pdf(_RESULT)
        assert isinstance(data, bytes)
        assert data[:5] == b"%PDF-"
        assert len(data) > 1000

    def test_font_embedded(self):
        # шрифт с кириллицей встроен (иначе русский текст — квадраты)
        data = plan_to_pdf(_RESULT)
        assert b"FinpilotSans" in data or b"Liberation" in data

    def test_empty_result_does_not_crash(self):
        assert plan_to_pdf({})[:5] == b"%PDF-"
